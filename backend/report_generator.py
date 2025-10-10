"""
Enhanced report generator with rich PDFs, charts, and templates
"""
import os
import json
import logging
from datetime import datetime, timedelta
from typing import Dict, Any, List, Optional, Tuple
from dataclasses import dataclass
from enum import Enum
import io
import base64

# PDF generation
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
from reportlab.graphics.shapes import Drawing, Rect, String
from reportlab.graphics.charts.barcharts import VerticalBarChart
from reportlab.graphics.charts.linecharts import HorizontalLineChart
from reportlab.graphics.charts.piecharts import Pie
from reportlab.graphics import renderPDF

# Excel generation
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.drawing.image import Image as ExcelImage

# Chart generation
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
from matplotlib.backends.backend_agg import FigureCanvasAgg
import seaborn as sns

logger = logging.getLogger(__name__)

class ReportFormat(Enum):
    PDF = "pdf"
    EXCEL = "xlsx"
    CSV = "csv"
    JSON = "json"

class ChartType(Enum):
    BAR = "bar"
    LINE = "line"
    PIE = "pie"
    SCATTER = "scatter"
    HEATMAP = "heatmap"

@dataclass
class ChartData:
    title: str
    chart_type: ChartType
    data: Dict[str, Any]
    x_label: str = ""
    y_label: str = ""
    colors: List[str] = None

@dataclass
class ReportTemplate:
    id: str
    name: str
    description: str
    sections: List[Dict[str, Any]]
    default_format: ReportFormat
    chart_configs: List[Dict[str, Any]]
    styling: Dict[str, Any]

class ReportGenerator:
    """Enhanced report generator with rich formatting and charts"""
    
    def __init__(self):
        self.templates = self._load_templates()
        self.chart_colors = [
            '#1f77b4', '#ff7f0e', '#2ca02c', '#d62728', '#9467bd',
            '#8c564b', '#e377c2', '#7f7f7f', '#bcbd22', '#17becf'
        ]
    
    def _load_templates(self) -> Dict[str, ReportTemplate]:
        """Load report templates"""
        templates = {
            'faculty_analytics': ReportTemplate(
                id='faculty_analytics',
                name='Faculty Analytics Report',
                description='Comprehensive faculty performance analysis',
                sections=[
                    {'type': 'header', 'content': 'Faculty Performance Analytics'},
                    {'type': 'summary', 'content': 'Executive Summary'},
                    {'type': 'charts', 'content': 'Performance Charts'},
                    {'type': 'table', 'content': 'Detailed Faculty Data'},
                    {'type': 'recommendations', 'content': 'Recommendations'}
                ],
                default_format=ReportFormat.PDF,
                chart_configs=[
                    {'type': 'bar', 'title': 'Faculty Ratings', 'data_field': 'ratings'},
                    {'type': 'pie', 'title': 'Department Distribution', 'data_field': 'departments'},
                    {'type': 'line', 'title': 'Trend Analysis', 'data_field': 'trends'}
                ],
                styling={
                    'primary_color': '#2563eb',
                    'secondary_color': '#64748b',
                    'accent_color': '#f59e0b',
                    'font_family': 'Helvetica',
                    'font_size': 12
                }
            ),
            'student_feedback': ReportTemplate(
                id='student_feedback',
                name='Student Feedback Report',
                description='Student feedback analysis and insights',
                sections=[
                    {'type': 'header', 'content': 'Student Feedback Analysis'},
                    {'type': 'summary', 'content': 'Feedback Summary'},
                    {'type': 'charts', 'content': 'Feedback Charts'},
                    {'type': 'table', 'content': 'Detailed Feedback Data'},
                    {'type': 'insights', 'content': 'Key Insights'}
                ],
                default_format=ReportFormat.PDF,
                chart_configs=[
                    {'type': 'bar', 'title': 'Question Ratings', 'data_field': 'question_ratings'},
                    {'type': 'pie', 'title': 'Response Distribution', 'data_field': 'response_distribution'},
                    {'type': 'line', 'title': 'Feedback Trends', 'data_field': 'trends'}
                ],
                styling={
                    'primary_color': '#059669',
                    'secondary_color': '#6b7280',
                    'accent_color': '#dc2626',
                    'font_family': 'Helvetica',
                    'font_size': 12
                }
            ),
            'department_summary': ReportTemplate(
                id='department_summary',
                name='Department Summary Report',
                description='Department-wide performance summary',
                sections=[
                    {'type': 'header', 'content': 'Department Performance Summary'},
                    {'type': 'summary', 'content': 'Department Overview'},
                    {'type': 'charts', 'content': 'Performance Charts'},
                    {'type': 'table', 'content': 'Department Data'},
                    {'type': 'comparison', 'content': 'Cross-Department Comparison'}
                ],
                default_format=ReportFormat.PDF,
                chart_configs=[
                    {'type': 'bar', 'title': 'Department Ratings', 'data_field': 'department_ratings'},
                    {'type': 'heatmap', 'title': 'Performance Heatmap', 'data_field': 'performance_matrix'},
                    {'type': 'scatter', 'title': 'Rating vs Response Count', 'data_field': 'rating_vs_responses'}
                ],
                styling={
                    'primary_color': '#7c3aed',
                    'secondary_color': '#9ca3af',
                    'accent_color': '#f97316',
                    'font_family': 'Helvetica',
                    'font_size': 12
                }
            )
        }
        return templates
    
    async def generate_report(
        self,
        template_id: str,
        data: Dict[str, Any],
        format: ReportFormat = None,
        custom_sections: Optional[List[Dict[str, Any]]] = None
    ) -> bytes:
        """Generate a report using the specified template"""
        try:
            template = self.templates.get(template_id)
            if not template:
                raise ValueError(f"Template not found: {template_id}")
            
            format = format or template.default_format
            
            if format == ReportFormat.PDF:
                return await self._generate_pdf_report(template, data, custom_sections)
            elif format == ReportFormat.EXCEL:
                return await self._generate_excel_report(template, data, custom_sections)
            elif format == ReportFormat.CSV:
                return await self._generate_csv_report(template, data, custom_sections)
            elif format == ReportFormat.JSON:
                return await self._generate_json_report(template, data, custom_sections)
            else:
                raise ValueError(f"Unsupported format: {format}")
                
        except Exception as e:
            logger.error(f"Report generation error: {e}")
            raise
    
    async def _generate_pdf_report(
        self,
        template: ReportTemplate,
        data: Dict[str, Any],
        custom_sections: Optional[List[Dict[str, Any]]] = None
    ) -> bytes:
        """Generate PDF report with rich formatting and charts"""
        try:
            buffer = io.BytesIO()
            doc = SimpleDocTemplate(
                buffer,
                pagesize=A4,
                rightMargin=72,
                leftMargin=72,
                topMargin=72,
                bottomMargin=18
            )
            
            # Get styles
            styles = getSampleStyleSheet()
            custom_styles = self._create_custom_styles(template.styling)
            
            # Build story (content)
            story = []
            
            # Add sections
            sections = custom_sections or template.sections
            for section in sections:
                section_content = await self._build_section_content(
                    section, template, data, custom_styles
                )
                story.extend(section_content)
            
            # Build PDF
            doc.build(story)
            buffer.seek(0)
            return buffer.getvalue()
            
        except Exception as e:
            logger.error(f"PDF generation error: {e}")
            raise
    
    async def _generate_excel_report(
        self,
        template: ReportTemplate,
        data: Dict[str, Any],
        custom_sections: Optional[List[Dict[str, Any]]] = None
    ) -> bytes:
        """Generate Excel report with charts and formatting"""
        try:
            workbook = openpyxl.Workbook()
            
            # Remove default sheet
            workbook.remove(workbook.active)
            
            # Add sections as worksheets
            sections = custom_sections or template.sections
            for i, section in enumerate(sections):
                sheet_name = section.get('content', f'Section {i+1}')[:31]  # Excel sheet name limit
                worksheet = workbook.create_sheet(sheet_name)
                
                # Add section content
                await self._build_excel_section(worksheet, section, template, data)
            
            # Save to bytes
            buffer = io.BytesIO()
            workbook.save(buffer)
            buffer.seek(0)
            return buffer.getvalue()
            
        except Exception as e:
            logger.error(f"Excel generation error: {e}")
            raise
    
    async def _generate_csv_report(
        self,
        template: ReportTemplate,
        data: Dict[str, Any],
        custom_sections: Optional[List[Dict[str, Any]]] = None
    ) -> bytes:
        """Generate CSV report"""
        try:
            import csv
            
            buffer = io.StringIO()
            writer = csv.writer(buffer)
            
            # Add header
            writer.writerow([template.name])
            writer.writerow([f"Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S')}"])
            writer.writerow([])
            
            # Add data sections
            sections = custom_sections or template.sections
            for section in sections:
                if section['type'] == 'table':
                    table_data = self._extract_table_data(section, data)
                    for row in table_data:
                        writer.writerow(row)
                    writer.writerow([])
            
            return buffer.getvalue().encode('utf-8')
            
        except Exception as e:
            logger.error(f"CSV generation error: {e}")
            raise
    
    async def _generate_json_report(
        self,
        template: ReportTemplate,
        data: Dict[str, Any],
        custom_sections: Optional[List[Dict[str, Any]]] = None
    ) -> bytes:
        """Generate JSON report"""
        try:
            report_data = {
                'template': template.id,
                'generated_at': datetime.utcnow().isoformat(),
                'sections': custom_sections or template.sections,
                'data': data,
                'metadata': {
                    'format': 'json',
                    'version': '1.0'
                }
            }
            
            return json.dumps(report_data, indent=2, default=str).encode('utf-8')
            
        except Exception as e:
            logger.error(f"JSON generation error: {e}")
            raise
    
    async def _build_section_content(
        self,
        section: Dict[str, Any],
        template: ReportTemplate,
        data: Dict[str, Any],
        styles: Dict[str, Any]
    ) -> List:
        """Build content for a PDF section"""
        content = []
        section_type = section['type']
        
        if section_type == 'header':
            content.append(Paragraph(section['content'], styles['title']))
            content.append(Spacer(1, 12))
            
        elif section_type == 'summary':
            content.append(Paragraph(section['content'], styles['heading']))
            content.append(Spacer(1, 6))
            
            # Add summary data
            summary_data = self._extract_summary_data(data)
            for key, value in summary_data.items():
                content.append(Paragraph(f"<b>{key}:</b> {value}", styles['normal']))
            content.append(Spacer(1, 12))
            
        elif section_type == 'charts':
            content.append(Paragraph(section['content'], styles['heading']))
            content.append(Spacer(1, 6))
            
            # Add charts
            for chart_config in template.chart_configs:
                chart_data = self._extract_chart_data(chart_config, data)
                if chart_data:
                    chart_image = await self._create_chart_image(chart_data)
                    if chart_image:
                        content.append(Image(chart_image, width=6*inch, height=4*inch))
                        content.append(Spacer(1, 12))
            
        elif section_type == 'table':
            content.append(Paragraph(section['content'], styles['heading']))
            content.append(Spacer(1, 6))
            
            # Add table
            table_data = self._extract_table_data(section, data)
            if table_data:
                table = Table(table_data)
                table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 14),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black)
                ]))
                content.append(table)
                content.append(Spacer(1, 12))
        
        return content
    
    async def _build_excel_section(
        self,
        worksheet,
        section: Dict[str, Any],
        template: ReportTemplate,
        data: Dict[str, Any]
    ):
        """Build content for an Excel section"""
        section_type = section['type']
        
        if section_type == 'header':
            worksheet['A1'] = section['content']
            worksheet['A1'].font = Font(size=16, bold=True)
            worksheet['A1'].alignment = Alignment(horizontal='center')
            
        elif section_type == 'summary':
            worksheet['A2'] = section['content']
            worksheet['A2'].font = Font(size=14, bold=True)
            
            # Add summary data
            summary_data = self._extract_summary_data(data)
            row = 3
            for key, value in summary_data.items():
                worksheet[f'A{row}'] = f"{key}: {value}"
                row += 1
                
        elif section_type == 'charts':
            worksheet['A1'] = section['content']
            worksheet['A1'].font = Font(size=14, bold=True)
            
            # Add charts
            for chart_config in template.chart_configs:
                chart_data = self._extract_chart_data(chart_config, data)
                if chart_data:
                    await self._create_excel_chart(worksheet, chart_data)
                    
        elif section_type == 'table':
            worksheet['A1'] = section['content']
            worksheet['A1'].font = Font(size=14, bold=True)
            
            # Add table
            table_data = self._extract_table_data(section, data)
            if table_data:
                for row_idx, row in enumerate(table_data, start=2):
                    for col_idx, cell in enumerate(row, start=1):
                        worksheet.cell(row=row_idx, column=col_idx, value=cell)
                        
                        # Style header row
                        if row_idx == 2:
                            worksheet.cell(row=row_idx, column=col_idx).font = Font(bold=True)
                            worksheet.cell(row=row_idx, column=col_idx).fill = PatternFill(
                                start_color='CCCCCC', end_color='CCCCCC', fill_type='solid'
                            )
    
    def _create_custom_styles(self, styling: Dict[str, Any]) -> Dict[str, Any]:
        """Create custom styles for PDF"""
        styles = getSampleStyleSheet()
        
        custom_styles = {
            'title': ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=24,
                spaceAfter=30,
                alignment=TA_CENTER,
                textColor=colors.HexColor(styling['primary_color'])
            ),
            'heading': ParagraphStyle(
                'CustomHeading',
                parent=styles['Heading2'],
                fontSize=16,
                spaceAfter=12,
                textColor=colors.HexColor(styling['primary_color'])
            ),
            'normal': ParagraphStyle(
                'CustomNormal',
                parent=styles['Normal'],
                fontSize=styling['font_size'],
                spaceAfter=6
            )
        }
        
        return custom_styles
    
    def _extract_summary_data(self, data: Dict[str, Any]) -> Dict[str, str]:
        """Extract summary data from report data"""
        summary = {}
        
        if 'total_responses' in data:
            summary['Total Responses'] = str(data['total_responses'])
        if 'average_rating' in data:
            summary['Average Rating'] = f"{data['average_rating']:.2f}"
        if 'faculty_count' in data:
            summary['Faculty Count'] = str(data['faculty_count'])
        if 'department_count' in data:
            summary['Department Count'] = str(data['department_count'])
        
        return summary
    
    def _extract_chart_data(self, chart_config: Dict[str, Any], data: Dict[str, Any]) -> Optional[ChartData]:
        """Extract chart data from report data"""
        data_field = chart_config.get('data_field')
        if data_field not in data:
            return None
        
        chart_data = data[data_field]
        if not chart_data:
            return None
        
        return ChartData(
            title=chart_config['title'],
            chart_type=ChartType(chart_config['type']),
            data=chart_data,
            colors=self.chart_colors
        )
    
    def _extract_table_data(self, section: Dict[str, Any], data: Dict[str, Any]) -> List[List[str]]:
        """Extract table data from report data"""
        # This would be implemented based on the specific data structure
        # For now, return a sample table
        return [
            ['Faculty', 'Rating', 'Responses', 'Department'],
            ['John Doe', '8.5', '45', 'Computer Science'],
            ['Jane Smith', '7.8', '38', 'Mathematics'],
            ['Bob Johnson', '9.2', '52', 'Physics']
        ]
    
    async def _create_chart_image(self, chart_data: ChartData) -> Optional[bytes]:
        """Create chart image for PDF"""
        try:
            plt.style.use('seaborn-v0_8')
            fig, ax = plt.subplots(figsize=(10, 6))
            
            if chart_data.chart_type == ChartType.BAR:
                self._create_bar_chart(ax, chart_data)
            elif chart_data.chart_type == ChartType.LINE:
                self._create_line_chart(ax, chart_data)
            elif chart_data.chart_type == ChartType.PIE:
                self._create_pie_chart(ax, chart_data)
            
            ax.set_title(chart_data.title, fontsize=14, fontweight='bold')
            ax.set_xlabel(chart_data.x_label)
            ax.set_ylabel(chart_data.y_label)
            
            # Save to bytes
            buffer = io.BytesIO()
            plt.savefig(buffer, format='png', dpi=300, bbox_inches='tight')
            buffer.seek(0)
            plt.close(fig)
            
            return buffer.getvalue()
            
        except Exception as e:
            logger.error(f"Chart creation error: {e}")
            return None
    
    async def _create_excel_chart(self, worksheet, chart_data: ChartData):
        """Create chart in Excel worksheet"""
        try:
            if chart_data.chart_type == ChartType.BAR:
                chart = BarChart()
                chart.title = chart_data.title
                chart.x_axis.title = chart_data.x_label
                chart.y_axis.title = chart_data.y_label
                
                # Add chart to worksheet
                worksheet.add_chart(chart, "E2")
                
        except Exception as e:
            logger.error(f"Excel chart creation error: {e}")
    
    def _create_bar_chart(self, ax, chart_data: ChartData):
        """Create bar chart"""
        labels = list(chart_data.data.keys())
        values = list(chart_data.data.values())
        
        bars = ax.bar(labels, values, color=chart_data.colors[:len(labels)])
        ax.set_xticklabels(labels, rotation=45, ha='right')
        
        # Add value labels on bars
        for bar, value in zip(bars, values):
            ax.text(bar.get_x() + bar.get_width()/2, bar.get_height() + 0.1,
                   f'{value:.1f}', ha='center', va='bottom')
    
    def _create_line_chart(self, ax, chart_data: ChartData):
        """Create line chart"""
        labels = list(chart_data.data.keys())
        values = list(chart_data.data.values())
        
        ax.plot(labels, values, marker='o', linewidth=2, markersize=6)
        ax.set_xticklabels(labels, rotation=45, ha='right')
        ax.grid(True, alpha=0.3)
    
    def _create_pie_chart(self, ax, chart_data: ChartData):
        """Create pie chart"""
        labels = list(chart_data.data.keys())
        values = list(chart_data.data.values())
        
        wedges, texts, autotexts = ax.pie(values, labels=labels, autopct='%1.1f%%',
                                         colors=chart_data.colors[:len(labels)])
        
        # Customize text
        for autotext in autotexts:
            autotext.set_color('white')
            autotext.set_fontweight('bold')
