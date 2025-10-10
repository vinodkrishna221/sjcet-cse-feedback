import pandas as pd
import io
import base64
from typing import List, Dict, Any, Tuple
from datetime import datetime
import logging
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

logger = logging.getLogger(__name__)

def calculate_letter_grade(weighted_score: float) -> str:
    """Convert weighted score to letter grade"""
    if weighted_score >= 95:
        return "A+"
    elif weighted_score >= 90:
        return "A"
    elif weighted_score >= 85:
        return "B+"
    elif weighted_score >= 80:
        return "B"
    elif weighted_score >= 70:
        return "C"
    elif weighted_score >= 60:
        return "D"
    else:
        return "F"

def format_report_data(feedback_data: List[Dict[str, Any]]) -> Tuple[List[Dict[str, Any]], Dict[str, Any]]:
    """Structure data with faculty in rows, questions in columns"""
    from models import FEEDBACK_QUESTIONS
    
    # Group by faculty_id and subject
    faculty_groups = {}
    for item in feedback_data:
        key = f"{item['faculty_id']}_{item['subject']}"
        if key not in faculty_groups:
            faculty_groups[key] = {
                'faculty_id': item['faculty_id'],
                'faculty_name': item['faculty_name'],
                'subject': item['subject'],
                'question_ratings': {},
                'weighted_scores': [],
                'total_feedback': 0
            }
        
        faculty_groups[key]['total_feedback'] += item['total_feedback']
        faculty_groups[key]['weighted_scores'].extend(item['weighted_scores'])
        
        # Aggregate question ratings
        for q_id, rating in item['question_wise_ratings'].items():
            if q_id not in faculty_groups[key]['question_ratings']:
                faculty_groups[key]['question_ratings'][q_id] = []
            faculty_groups[key]['question_ratings'][q_id].append(rating)
    
    # Calculate averages and format data
    report_rows = []
    total_faculty = len(faculty_groups)
    total_feedback = 0
    all_weighted_scores = []
    
    for faculty_data in faculty_groups.values():
        # Calculate average weighted score
        avg_weighted_score = sum(faculty_data['weighted_scores']) / len(faculty_data['weighted_scores']) if faculty_data['weighted_scores'] else 0
        letter_grade = calculate_letter_grade(avg_weighted_score)
        
        # Calculate question averages
        question_averages = {}
        for q_id, ratings in faculty_data['question_ratings'].items():
            question_averages[q_id] = sum(ratings) / len(ratings) if ratings else 0
        
        # Create row data
        row_data = {
            'Faculty Name': faculty_data['faculty_name'],
            'Subject': faculty_data['subject'],
            'Total Feedback Count': faculty_data['total_feedback'],
            'Overall Weighted Score (%)': round(avg_weighted_score, 2),
            'Letter Grade': letter_grade
        }
        
        # Add question averages
        for question in FEEDBACK_QUESTIONS:
            q_id = question['id']
            avg_rating = question_averages.get(q_id, 0)
            row_data[f"{question['question']} (Avg)"] = round(avg_rating, 2)
        
        report_rows.append(row_data)
        total_feedback += faculty_data['total_feedback']
        all_weighted_scores.append(avg_weighted_score)
    
    # Calculate section summary
    section_avg = sum(all_weighted_scores) / len(all_weighted_scores) if all_weighted_scores else 0
    highest_score = max(all_weighted_scores) if all_weighted_scores else 0
    lowest_score = min(all_weighted_scores) if all_weighted_scores else 0
    
    # Grade distribution
    grade_distribution = {}
    for score in all_weighted_scores:
        grade = calculate_letter_grade(score)
        grade_distribution[grade] = grade_distribution.get(grade, 0) + 1
    
    summary_metrics = {
        'total_faculty': total_faculty,
        'total_feedback': total_feedback,
        'section_average': round(section_avg, 2),
        'highest_score': round(highest_score, 2),
        'lowest_score': round(lowest_score, 2),
        'grade_distribution': grade_distribution
    }
    
    return report_rows, summary_metrics

def generate_csv_report(report_data: List[Dict[str, Any]], summary_metrics: Dict[str, Any], 
                       department: str, batch_year: str, section: str) -> bytes:
    """Generate CSV report"""
    try:
        # Create DataFrame
        df = pd.DataFrame(report_data)
        
        # Add summary row
        summary_row = {
            'Faculty Name': 'SECTION SUMMARY',
            'Subject': '',
            'Total Feedback Count': summary_metrics['total_feedback'],
            'Overall Weighted Score (%)': summary_metrics['section_average'],
            'Letter Grade': calculate_letter_grade(summary_metrics['section_average'])
        }
        
        # Add question columns for summary (empty)
        for col in df.columns:
            if col not in summary_row:
                summary_row[col] = ''
        
        # Append summary row
        df = pd.concat([df, pd.DataFrame([summary_row])], ignore_index=True)
        
        # Convert to CSV
        csv_buffer = io.StringIO()
        df.to_csv(csv_buffer, index=False)
        csv_content = csv_buffer.getvalue()
        csv_buffer.close()
        
        return csv_content.encode('utf-8')
        
    except Exception as e:
        logger.error(f"Error generating CSV report: {e}")
        raise

def generate_excel_report(report_data: List[Dict[str, Any]], summary_metrics: Dict[str, Any],
                         department: str, batch_year: str, section: str) -> bytes:
    """Generate Excel report with formatting"""
    try:
        # Create DataFrame
        df = pd.DataFrame(report_data)
        
        # Create workbook and worksheet
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"Faculty Report - {section}"
        
        # Add header
        ws['A1'] = f"Faculty Feedback Report - {department} - {batch_year} - Section {section}"
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:H1')
        
        # Add generation info
        ws['A2'] = f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws['A2'].font = Font(size=10)
        
        # Add data starting from row 4
        start_row = 4
        
        # Headers
        headers = list(df.columns)
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=start_row, column=col_idx, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # Data rows
        for row_idx, row_data in enumerate(report_data, start_row + 1):
            for col_idx, header in enumerate(headers, 1):
                value = row_data.get(header, '')
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                
                # Format numeric columns
                if 'Score' in header or 'Avg' in header:
                    cell.number_format = '0.00'
                elif 'Count' in header:
                    cell.number_format = '0'
                
                # Highlight letter grades
                if header == 'Letter Grade':
                    if value == 'A+':
                        cell.fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
                    elif value in ['A', 'B+']:
                        cell.fill = PatternFill(start_color="98FB98", end_color="98FB98", fill_type="solid")
                    elif value in ['B', 'C']:
                        cell.fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
                    elif value in ['D', 'F']:
                        cell.fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")
        
        # Add summary section
        summary_start_row = start_row + len(report_data) + 2
        ws.cell(row=summary_start_row, column=1, value="SECTION SUMMARY").font = Font(bold=True, size=12)
        
        summary_data = [
            ("Total Faculty", summary_metrics['total_faculty']),
            ("Total Feedback", summary_metrics['total_feedback']),
            ("Section Average", f"{summary_metrics['section_average']}%"),
            ("Highest Score", f"{summary_metrics['highest_score']}%"),
            ("Lowest Score", f"{summary_metrics['lowest_score']}%")
        ]
        
        for idx, (label, value) in enumerate(summary_data):
            ws.cell(row=summary_start_row + 1 + idx, column=1, value=label).font = Font(bold=True)
            ws.cell(row=summary_start_row + 1 + idx, column=2, value=value)
        
        # Add grade distribution
        grade_start_row = summary_start_row + len(summary_data) + 2
        ws.cell(row=grade_start_row, column=1, value="GRADE DISTRIBUTION").font = Font(bold=True, size=12)
        
        for idx, (grade, count) in enumerate(summary_metrics['grade_distribution'].items()):
            ws.cell(row=grade_start_row + 1 + idx, column=1, value=grade).font = Font(bold=True)
            ws.cell(row=grade_start_row + 1 + idx, column=2, value=count)
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to bytes
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_content = excel_buffer.getvalue()
        excel_buffer.close()
        
        return excel_content
        
    except Exception as e:
        logger.error(f"Error generating Excel report: {e}")
        raise

def generate_pdf_report(report_data: List[Dict[str, Any]], summary_metrics: Dict[str, Any],
                       department: str, batch_year: str, section: str) -> bytes:
    """Generate PDF report"""
    try:
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=1*inch)
        
        # Get styles
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=16,
            spaceAfter=30,
            alignment=1  # Center alignment
        )
        
        # Content
        story = []
        
        # Title
        title = f"Faculty Feedback Report<br/>{department} - {batch_year} - Section {section}"
        story.append(Paragraph(title, title_style))
        
        # Generation info
        gen_info = f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        story.append(Paragraph(gen_info, styles['Normal']))
        story.append(Spacer(1, 20))
        
        # Prepare table data
        if report_data:
            headers = list(report_data[0].keys())
            table_data = [headers]
            
            for row in report_data:
                table_data.append([str(row.get(header, '')) for header in headers])
            
            # Create table
            table = Table(table_data)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 8),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('FONTSIZE', (0, 1), (-1, -1), 7),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ]))
            
            story.append(table)
            story.append(Spacer(1, 20))
        
        # Summary section
        story.append(Paragraph("SECTION SUMMARY", styles['Heading2']))
        
        summary_text = f"""
        <b>Total Faculty:</b> {summary_metrics['total_faculty']}<br/>
        <b>Total Feedback:</b> {summary_metrics['total_feedback']}<br/>
        <b>Section Average:</b> {summary_metrics['section_average']}%<br/>
        <b>Highest Score:</b> {summary_metrics['highest_score']}%<br/>
        <b>Lowest Score:</b> {summary_metrics['lowest_score']}%<br/>
        """
        story.append(Paragraph(summary_text, styles['Normal']))
        
        # Grade distribution
        story.append(Spacer(1, 10))
        story.append(Paragraph("GRADE DISTRIBUTION", styles['Heading2']))
        
        grade_text = "<br/>".join([f"<b>{grade}:</b> {count}" for grade, count in summary_metrics['grade_distribution'].items()])
        story.append(Paragraph(grade_text, styles['Normal']))
        
        # Build PDF
        doc.build(story)
        
        pdf_content = buffer.getvalue()
        buffer.close()
        
        return pdf_content
        
    except Exception as e:
        logger.error(f"Error generating PDF report: {e}")
        raise
